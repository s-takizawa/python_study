import openpyxl

print('start')

wb = openpyxl.load_workbook(r'input_data\data1.xlsx', data_only=True)
print(type(wb))
print(wb.sheetnames)

# wb.save(r'input_data\hoge.xlsx')
ws = wb.active
print(ws)

sheet_1 = wb['sheet_1']

print(sheet_1['B5'].value)
print(sheet_1['C5'].value)
print(sheet_1['D5'].value)

c = sheet_1['C4']
msg = 'Row %s, Column %s is %s' % (c.row, c.column, c.value)
print(msg)
msg = 'Cell %s is %s' % (c.coordinate, c.value)
print(msg)

print('max_row:' + str(sheet_1.max_row))
print('max_column:' + str(sheet_1.max_column))
print(tuple(sheet_1['A1' : 'E5']))

print('-------')

row = int(sheet_1['B1'].value)  # 行数
col = int(sheet_1['D1'].value)  # 列数
print('col={0:d},row={1:d}'.format(col, row))

for r in range(3, row + 3):
    print(sheet_1.cell(row=r, column=3).value)


print('end')
